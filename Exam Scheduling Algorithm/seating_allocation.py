"""
Seating Arrangement Allocation System
--------------------------------------
This system allocates students to exam halls with the following features:
1. Linear department seating allocation
2. Teacher/Invigilator assignment
3. PDF generation (Student & Faculty versions)
4. Excel output with comprehensive reports
"""

import pandas as pd
import numpy as np
import random
import os
import sqlite3
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.backends.backend_pdf import PdfPages
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
try:
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    PatternFill = Font = Border = Side = Alignment = get_column_letter = None

# Database path - shared with exam scheduling
DB_PATH = os.path.join(os.path.dirname(__file__), '..', 'Exam Scheduling Algorithm', 'exam_scheduling.db')


class SeatingAllocationSystem:
    def __init__(self, halls_file=None, students_file=None, teachers_file=None, session='FN', exam_type='Internal', year=1, internal_number=1, selected_halls=None, selected_teachers=None, use_database=True, exam_date=None):
        """Initialize the seating allocation system
        
        Args:
            use_database: If True, load data from database. If False, load from CSV files.
            exam_date: For SEM exams, the specific date+session to allocate seats for (DD.MM.YYYY)
        """
        self.exam_type = exam_type
        self.exam_date = exam_date
        self.session = session
        self.year = year
        
        if use_database:
            # Load data from database
            self._load_from_database(year, selected_halls, selected_teachers)
        else:
            # Load data from CSV files (legacy mode)
            # Read halls data with columns information
            self.halls_df = pd.read_csv(halls_file)
            self.halls_df.columns = self.halls_df.columns.str.strip()
            
            # Read students data - preserve register numbers as strings
            self.students_df = pd.read_csv(students_file, dtype={'Register Number': str})
            self.students_df.columns = self.students_df.columns.str.strip()
            
            # Read teachers data
            self.teachers_df = pd.read_csv(teachers_file)
            self.teachers_df.columns = self.teachers_df.columns.str.strip()
            
            # Filter halls if specific halls are selected
            if selected_halls:
                self.halls_df = self.halls_df[self.halls_df['hallno'].isin(selected_halls)].reset_index(drop=True)
        
        # Filter teachers if specific teachers are selected
        if selected_teachers:
            self.teachers_df = self.teachers_df[self.teachers_df['Name'].isin(selected_teachers)].reset_index(drop=True)
        
        # Prepare data structures
        self.allocations = []
        self.hall_wise_allocations = {}
        self.teacher_assignments = {}
        self.session = session  # 'FN' or 'AN'
        self.exam_type = exam_type  # 'Internal' or 'SEM'
        self.year = year  # Academic year (1, 2, 3, or 4)
        self.internal_number = internal_number  # 1 or 2 (only for Internal exams)
        self.generation_date = datetime.now().strftime('%Y-%m-%d')
    
    def _load_from_database(self, year, selected_halls=None, selected_teachers=None):
        """Load data from shared database"""
        import json
        conn = sqlite3.connect(DB_PATH)
        
        # Load halls data
        halls_query = "SELECT hall_name as hallno, capacity, columns FROM halls WHERE active = 1"
        self.halls_df = pd.read_sql_query(halls_query, conn)
        
        # Filter halls if selected
        if selected_halls:
            self.halls_df = self.halls_df[self.halls_df['hallno'].isin(selected_halls)].reset_index(drop=True)
        
        # Load students based on exam type
        if self.exam_type == 'SEMESTER' and self.exam_date:
            # For SEM exams, get students based on scheduled subjects for this date+session
            students_query = '''
                SELECT DISTINCT 
                    s.reg_no as "Register Number", 
                    s.name as "Name", 
                    s.department as "Department",
                    s.year as "Student Year",
                    s.arrears as "Arrears"
                FROM students s
                JOIN student_subjects ss ON s.student_id = ss.student_id
                JOIN subjects sub ON ss.subject_id = sub.subject_id
                JOIN schedules sch ON sub.subject_id = sch.subject_id
                WHERE sch.exam_date = ? AND sch.session = ? AND s.active = 1
                ORDER BY s.department, s.reg_no
            '''
            self.students_df = pd.read_sql_query(students_query, conn, params=(self.exam_date, self.session))
            
            # Filter students: include regular students + arrear students (subject_code in arrears array)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT DISTINCT sub.subject_code
                FROM schedules sch
                JOIN subjects sub ON sch.subject_id = sub.subject_id
                WHERE sch.exam_date = ? AND sch.session = ?
            ''', (self.exam_date, self.session))
            
            scheduled_subjects = [row[0] for row in cursor.fetchall()]
            
            # Filter students who have these subjects (regular or arrear)
            filtered_students = []
            for _, row in self.students_df.iterrows():
                arrears = json.loads(row['Arrears']) if row['Arrears'] else []
                # Include if any scheduled subject is in their arrears array
                if any(sub_code in arrears for sub_code in scheduled_subjects):
                    filtered_students.append(row)
                # Also include regular students (those whose year matches scheduled subjects)
                else:
                    # Check if they're regular students for these subjects
                    cursor.execute('''
                        SELECT COUNT(*) FROM student_subjects ss
                        JOIN subjects sub ON ss.subject_id = sub.subject_id
                        JOIN students st ON ss.student_id = st.student_id
                        WHERE st.reg_no = ? AND sub.subject_code IN ({})
                              AND ss.is_arrear = 0
                    '''.format(','.join('?' * len(scheduled_subjects))), 
                    (row['Register Number'], *scheduled_subjects))
                    
                    if cursor.fetchone()[0] > 0:
                        filtered_students.append(row)
            
            self.students_df = pd.DataFrame(filtered_students)
            
        elif self.exam_type == 'Internal' and self.exam_date:
            # For Internal exams, get students enrolled in subjects for this session
            # NO ARREAR checking - only students enrolled in subjects scheduled for this session
            students_query = '''
                SELECT DISTINCT 
                    s.student_id,
                    s.reg_no as "Register Number", 
                    s.name as "Name", 
                    s.department as "Department",
                    s.year as "Student Year"
                FROM students s
                JOIN student_subjects ss ON s.student_id = ss.student_id
                JOIN schedules sch ON ss.subject_id = sch.subject_id
                WHERE sch.session = ? AND s.year = ? AND s.active = 1
                ORDER BY s.department, s.reg_no
            '''
            self.students_df = pd.read_sql_query(students_query, conn, params=(self.session, year))
            
        else:
            # Fallback: Load all students for the year (old internal logic)
            students_query = '''
                SELECT DISTINCT 
                    s.reg_no as "Register Number", 
                    s.name as "Name", 
                    s.department as "Department",
                    s.year as "Student Year"
                FROM students s
                WHERE s.year = ? AND s.active = 1
                ORDER BY s.department, s.reg_no
            '''
            self.students_df = pd.read_sql_query(students_query, conn, params=(year,))
        
        # Load teachers data
        teachers_query = "SELECT teacher_name as Name, department as Department FROM teachers WHERE active = 1"
        self.teachers_df = pd.read_sql_query(teachers_query, conn)
        
        # Filter teachers if selected
        if selected_teachers:
            self.teachers_df = self.teachers_df[self.teachers_df['Name'].isin(selected_teachers)].reset_index(drop=True)
        
        conn.close()
        
    def optimize_hall_selection(self):
        """
        Optimize hall selection to minimize empty spaces
        Returns list of hall indices sorted by efficiency
        """
        total_students = len(self.students_df)
        
        if self.exam_type == 'Internal':
            # For internal, capacity is benches, each bench holds 2 students
            self.halls_df['effective_capacity'] = self.halls_df['capacity'] * 2
        else:
            # For SEM, capacity is seats (1 student per bench)
            self.halls_df['effective_capacity'] = self.halls_df['capacity']
        
        # Sort halls by capacity (ascending) to pack efficiently
        halls_sorted = self.halls_df.sort_values('effective_capacity').reset_index(drop=True)
        
        # Greedy algorithm: select halls that minimize total waste
        selected_indices = []
        accumulated_capacity = 0
        
        for idx, hall in halls_sorted.iterrows():
            if accumulated_capacity >= total_students:
                break
            selected_indices.append(idx)
            accumulated_capacity += hall['effective_capacity']
        
        # If we need all halls, return all
        if accumulated_capacity < total_students:
            selected_indices = list(range(len(self.halls_df)))
        
        return selected_indices
    
    def allocate_seats_mixed_department(self):
        """
        Allocate seats based on exam type:
        - SEM: Linear allocation by department, 1 student per bench
        - Internal: Alternating departments, 2 students per bench from different departments
        """
        print("=" * 60)
        print(f"SEATING ALLOCATION - {self.exam_type.upper()} EXAM FORMAT")
        print("=" * 60)
        
        # Optimize hall selection to minimize waste
        optimal_halls = self.optimize_hall_selection()
        
        if self.exam_type in ['SEM', 'SEMESTER']:
            # Semester exam: Linear allocation, 1 student per bench
            allocations = self._allocate_sem_linear_optimized(optimal_halls)
        else:
            # Internal exam: Alternating departments, 2 students per bench
            allocations = self._allocate_internal_alternating_optimized(optimal_halls)
        
        self.allocations = pd.DataFrame(allocations)
        print(f"\nTotal students allocated: {len(self.allocations)}")
        
        # Create hall-wise summary
        self._create_hall_wise_summary()
        
        return self.allocations
    
    def _allocate_sem_linear_optimized(self, optimal_hall_indices):
        """Allocate for SEM exam: 1 student per bench with randomization and min 2 depts per hall"""
        # Group students by department and shuffle within each department
        departments = sorted(self.students_df['Department'].unique())
        dept_groups = {}
        
        for dept in departments:
            dept_students = self.students_df[self.students_df['Department'] == dept].copy()
            dept_students = dept_students.sort_values('Register Number').reset_index(drop=True)
            # Shuffle to add randomness
            dept_students = dept_students.sample(frac=1, random_state=42).reset_index(drop=True)
            dept_groups[dept] = dept_students
        
        # Create pointers for each department
        dept_pointers = {dept: 0 for dept in departments}
        
        allocations = []
        current_hall_position = 0
        current_seat_in_hall = 1
        total_students = len(self.students_df)
        total_allocated = 0
        
        # Track departments used in current hall
        current_hall_depts = set()
        hall_start_idx = 0
        
        while total_allocated < total_students and current_hall_position < len(optimal_hall_indices):
            current_hall_idx = optimal_hall_indices[current_hall_position]
            hall_no = self.halls_df.loc[current_hall_idx, 'hallno']
            hall_capacity = self.halls_df.loc[current_hall_idx, 'capacity']
            
            # Find available departments (prioritize ensuring min 2 depts per hall)
            available_depts = [dept for dept, ptr in dept_pointers.items() 
                             if ptr < len(dept_groups[dept])]
            
            if len(available_depts) == 0:
                break
            
            # Select department with controlled randomness
            # Ensure at least 2 different departments per hall
            if len(current_hall_depts) < 2:
                # Prefer departments not yet in this hall
                unused_depts = [d for d in available_depts if d not in current_hall_depts]
                if unused_depts:
                    selected_dept = random.choice(unused_depts)
                else:
                    selected_dept = random.choice(available_depts)
            else:
                # Random selection from all available
                selected_dept = random.choice(available_depts)
            
            current_hall_depts.add(selected_dept)
            
            student = dept_groups[selected_dept].iloc[dept_pointers[selected_dept]]
            
            # For SEM exams, each student gets their own bench
            # Seat numbers should be unique within each hall
            allocations.append({
                'Hall No': hall_no,
                'Seat No': current_seat_in_hall,
                'Register Number': student['Register Number'],
                'Name': student['Name'],
                'Department': student['Department'],
                'Bench Number': current_seat_in_hall  # Same as seat for SEM
            })
            
            dept_pointers[selected_dept] += 1
            total_allocated += 1
            current_seat_in_hall += 1
            
            # Move to next hall if current is full
            if current_seat_in_hall > hall_capacity:
                current_hall_position += 1
                current_seat_in_hall = 1
                current_hall_depts = set()
                hall_start_idx = len(allocations)
        
        # Print final hall info if not empty
        if current_hall_depts:
            print(f"  Hall {hall_no}: {len(current_hall_depts)} departments - {current_hall_depts}")
        
        print(f"Halls used: {current_hall_position + 1} out of {len(self.halls_df)}")
        return allocations
    
    def _allocate_sem_linear(self):
        """Wrapper for backward compatibility"""
        optimal_halls = list(range(len(self.halls_df)))
        return self._allocate_sem_linear_optimized(optimal_halls)
    
    def _allocate_internal_alternating_optimized(self, optimal_hall_indices):
        """Allocate for Internal exam: 2 students per bench with randomization and min 2 depts per hall"""
        # Group students by department and shuffle
        departments = sorted(self.students_df['Department'].unique())
        dept_groups = {}
        
        for dept in departments:
            dept_students = self.students_df[self.students_df['Department'] == dept].copy()
            dept_students = dept_students.sort_values('Register Number').reset_index(drop=True)
            # Shuffle for randomness
            dept_students = dept_students.sample(frac=1, random_state=42).reset_index(drop=True)
            dept_groups[dept] = dept_students
        
        # Create pointers for each department
        dept_pointers = {dept: 0 for dept in departments}
        
        allocations = []
        current_hall_position = 0
        current_seat_in_hall = 1
        total_students = len(self.students_df)
        total_allocated = 0
        
        # Track departments in current hall
        current_hall_depts = set()
        
        # For Internal exams, capacity represents benches
        while total_allocated < total_students and current_hall_position < len(optimal_hall_indices):
            current_hall_idx = optimal_hall_indices[current_hall_position]
            hall_no = self.halls_df.loc[current_hall_idx, 'hallno']
            hall_capacity = self.halls_df.loc[current_hall_idx, 'capacity']
            
            # Find available departments
            available_depts = [dept for dept, ptr in dept_pointers.items() 
                             if ptr < len(dept_groups[dept])]
            
            if len(available_depts) == 0:
                break
            
            # Select first student (ensure dept diversity in hall)
            if len(current_hall_depts) < 2:
                unused_depts = [d for d in available_depts if d not in current_hall_depts]
                dept1 = random.choice(unused_depts) if unused_depts else random.choice(available_depts)
            else:
                dept1 = random.choice(available_depts)
            
            current_hall_depts.add(dept1)
            student1 = dept_groups[dept1].iloc[dept_pointers[dept1]]
            
            allocations.append({
                'Hall No': hall_no,
                'Seat No': current_seat_in_hall,
                'Register Number': student1['Register Number'],
                'Name': student1['Name'],
                'Department': student1['Department']
            })
            
            dept_pointers[dept1] += 1
            total_allocated += 1
            
            # Try to allocate second student from different department (bench-mate)
            available_depts = [dept for dept, ptr in dept_pointers.items() 
                             if ptr < len(dept_groups[dept])]
            
            if len(available_depts) > 0:
                # Prefer different department for bench-mate
                other_depts = [d for d in available_depts if d != dept1]
                if other_depts:
                    dept2 = random.choice(other_depts)
                    current_hall_depts.add(dept2)
                    student2 = dept_groups[dept2].iloc[dept_pointers[dept2]]
                    
                    allocations.append({
                        'Hall No': hall_no,
                        'Seat No': current_seat_in_hall,  # Same seat for bench-mates
                        'Register Number': student2['Register Number'],
                        'Name': student2['Name'],
                        'Department': student2['Department']
                    })
                    
                    dept_pointers[dept2] += 1
                    total_allocated += 1
            
            current_seat_in_hall += 1
            
            # Move to next hall if current is full
            if current_seat_in_hall > hall_capacity:
                print(f"  Hall {hall_no}: {len(current_hall_depts)} departments - {current_hall_depts}")
                current_hall_position += 1
                current_seat_in_hall = 1
                current_hall_depts = set()
        
        # Print final hall info
        if current_hall_depts:
            print(f"  Hall {hall_no}: {len(current_hall_depts)} departments - {current_hall_depts}")
        
        print(f"Halls used: {current_hall_position + 1} out of {len(self.halls_df)}")
        print(f"Benches per hall: ~{hall_capacity}, Total capacity: ~{hall_capacity * 2} students")
        return allocations
    
    def _allocate_internal_alternating(self):
        """Wrapper for backward compatibility"""
        optimal_halls = list(range(len(self.halls_df)))
        return self._allocate_internal_alternating_optimized(optimal_halls)
    
    def allocate_seats_alternating_department(self):
        """
        Allocate seats with different departments alternating
        This ensures students from the same department don't sit next to each other
        """
        print("\n" + "=" * 60)
        print("SEATING ALLOCATION - ALTERNATING DEPARTMENT FORMAT")
        print("=" * 60)
        
        # Group students by department
        departments = self.students_df['Department'].unique()
        dept_groups = {dept: self.students_df[self.students_df['Department'] == dept].copy() 
                      for dept in departments}
        
        # Sort each department group by register number
        for dept in dept_groups:
            dept_groups[dept] = dept_groups[dept].sort_values('Register Number').reset_index(drop=True)
        
        # Create pointers for each department
        dept_pointers = {dept: 0 for dept in departments}
        dept_list = list(departments)
        
        # Allocate students to halls
        current_hall_idx = 0
        current_seat_in_hall = 1
        current_dept_idx = 0
        
        allocations = []
        
        total_allocated = 0
        total_students = len(self.students_df)
        
        while total_allocated < total_students:
            # Get current hall info
            hall_no = self.halls_df.loc[current_hall_idx, 'hallno']
            hall_capacity = self.halls_df.loc[current_hall_idx, 'capacity']
            
            # Try to allocate from current department
            dept = dept_list[current_dept_idx]
            dept_pointer = dept_pointers[dept]
            
            # Check if current department still has students
            if dept_pointer < len(dept_groups[dept]):
                student = dept_groups[dept].iloc[dept_pointer]
                
                allocations.append({
                    'Hall No': hall_no,
                    'Seat No': current_seat_in_hall,
                    'Register Number': student['Register Number'],
                    'Name': student['Name'],
                    'Department': student['Department']
                })
                
                dept_pointers[dept] += 1
                current_seat_in_hall += 1
                total_allocated += 1
                
            # Move to next department in rotation
            current_dept_idx = (current_dept_idx + 1) % len(dept_list)
            
            # Check if we need to move to next hall
            if current_seat_in_hall > hall_capacity:
                current_hall_idx += 1
                current_seat_in_hall = 1
                
                if current_hall_idx >= len(self.halls_df):
                    print("Warning: Ran out of halls!")
                    break
        
        allocations_df = pd.DataFrame(allocations)
        print(f"\nTotal students allocated: {total_allocated}")
        print(f"Halls used: {current_hall_idx + 1} out of {len(self.halls_df)}")
        
        return allocations_df
    
    def allocate_seats_same_department(self):
        """
        Allocate seats with same department students together
        Students from same department fill up halls before moving to next department
        """
        print("\n" + "=" * 60)
        print("SEATING ALLOCATION - SAME DEPARTMENT FORMAT")
        print("=" * 60)
        
        # Sort students by department and register number
        students_sorted = self.students_df.sort_values(
            ['Department', 'Register Number']
        ).reset_index(drop=True)
        
        allocations = []
        current_hall_idx = 0
        current_seat_in_hall = 1
        
        for idx, student in students_sorted.iterrows():
            hall_no = self.halls_df.loc[current_hall_idx, 'hallno']
            hall_capacity = self.halls_df.loc[current_hall_idx, 'capacity']
            
            allocations.append({
                'Hall No': hall_no,
                'Seat No': current_seat_in_hall,
                'Register Number': student['Register Number'],
                'Name': student['Name'],
                'Department': student['Department']
            })
            
            current_seat_in_hall += 1
            
            # Move to next hall if current hall is full
            if current_seat_in_hall > hall_capacity:
                current_hall_idx += 1
                current_seat_in_hall = 1
        
        allocations_df = pd.DataFrame(allocations)
        print(f"\nTotal students allocated: {len(allocations_df)}")
        print(f"Halls used: {current_hall_idx + 1} out of {len(self.halls_df)}")
        
        return allocations_df
    
    def _create_hall_wise_summary(self):
        """Create a summary of allocations by hall"""
        self.hall_wise_allocations = {}
        
        for hall_no in self.allocations['Hall No'].unique():
            hall_data = self.allocations[self.allocations['Hall No'] == hall_no].copy()
            hall_data = hall_data.sort_values('Seat No').reset_index(drop=True)
            self.hall_wise_allocations[hall_no] = hall_data
    
    def assign_teachers(self):
        """Assign teachers to halls (one-to-one assignment)"""
        print("\n" + "=" * 60)
        print("ASSIGNING TEACHERS TO HALLS")
        print("=" * 60)
        
        halls_used = sorted(self.hall_wise_allocations.keys())
        teachers_list = self.teachers_df['Name'].str.strip().tolist()
        
        # One-to-one assignment
        for idx, hall_no in enumerate(halls_used):
            if idx < len(teachers_list):
                self.teacher_assignments[hall_no] = teachers_list[idx]
            else:
                self.teacher_assignments[hall_no] = "To be assigned"
        
        print(f"\nAssigned {len(halls_used)} teachers to {len(halls_used)} halls")
        if len(teachers_list) > len(halls_used):
            print(f"Reserve teachers: {', '.join(teachers_list[len(halls_used):])}")
    
    def convert_to_2d_layout(self, hall_no):
        """Convert student list to 2D grid layout using hall-specific columns"""
        # Get the number of columns for this specific hall
        num_cols = self.halls_df[self.halls_df['hallno'] == hall_no]['columns'].values[0]
            
        hall_data = self.hall_wise_allocations[hall_no]
        hall_capacity = self.halls_df[self.halls_df['hallno'] == hall_no]['capacity'].values[0]
        
        if self.exam_type in ['SEM', 'SEMESTER']:
            # Semester Exam: 1 student per bench
            students = hall_data['Register Number'].tolist()
            num_rows = int(np.ceil(hall_capacity / num_cols))
            
            layout = []
            for row in range(num_rows):
                row_data = []
                for col in range(num_cols):
                    idx = row * num_cols + col
                    if idx < len(students):
                        row_data.append(students[idx])
                    else:
                        row_data.append("-")  # Empty seats shown as dash
                layout.append(row_data)
        
        else:  # Internal Exam
            # Internal Exam: 2 students per bench from different departments
            # Group by seat number to get bench-mates
            num_rows = int(np.ceil(hall_capacity / num_cols))
            
            layout = []
            bench_idx = 0
            
            for row in range(num_rows):
                row_data = []
                for col in range(num_cols):
                    bench_idx += 1
                    # Get students for this bench (same seat number)
                    bench_students = hall_data[hall_data['Seat No'] == bench_idx]
                    
                    if len(bench_students) == 0:
                        row_data.append({"left": "-", "right": "-"})
                    elif len(bench_students) == 1:
                        student = bench_students.iloc[0]
                        row_data.append({
                            "left": student['Register Number'],
                            "right": "-",  # Empty seat shown as dash
                            "dept_left": student['Department']
                        })
                    else:  # 2 students
                        student1 = bench_students.iloc[0]
                        student2 = bench_students.iloc[1]
                        row_data.append({
                            "left": student1['Register Number'],
                            "right": student2['Register Number'],
                            "dept_left": student1['Department'],
                            "dept_right": student2['Department']
                        })
                    
                layout.append(row_data)
        
        return layout, num_rows, num_cols
    
    def generate_hall_visual(self, hall_no, save_path=None):
        """Generate visual representation of hall layout using matplotlib"""
        layout, num_rows, num_cols = self.convert_to_2d_layout(hall_no)
        
        # Create figure in landscape orientation (11.69 x 8.27 inches = A4 landscape)
        fig, ax = plt.subplots(figsize=(11.69, 8.27))
        ax.axis('off')
        
        # Get hall info
        teacher = self.teacher_assignments.get(hall_no, "TBA")
        hall_capacity = self.halls_df[self.halls_df['hallno'] == hall_no]['capacity'].values[0]
        hall_data = self.hall_wise_allocations[hall_no]
        occupied = len(hall_data)
        
        # Get department breakdown
        dept_counts = hall_data['Department'].value_counts()
        dept_text = "\n".join([f"{dept}({count})" for dept, count in dept_counts.items()])
        
        # Add college header
        fig.text(0.5, 0.96, 'Marri Laxman Reddy Institute of Technology',
                ha='center', fontsize=16, fontweight='bold')
        fig.text(0.5, 0.93, 'Hyderabad - 43',
                ha='center', fontsize=11)
        fig.text(0.5, 0.90, '[An Autonomous Institution]',
                ha='center', fontsize=9, style='italic')
        
        # Add exam type to title
        if self.exam_type == 'Internal':
            roman_numeral = 'I' if self.internal_number == 1 else 'II'
            exam_type_text = f'Continuous Internal Assessment - {roman_numeral}'
        else:
            exam_type_text = 'End Semester Examination'
        fig.text(0.5, 0.87, f'SEATING ARRANGEMENT ({exam_type_text})',
                ha='center', fontsize=14, fontweight='bold')
        
        # Add date, session, and hall info
        # Use exam_date for SEM exams, current date for Internal exams
        if self.exam_type in ['SEM', 'SEMESTER'] and self.exam_date:
            # Convert from DD.MM.YYYY to DD-MM-YYYY format
            display_date = self.exam_date.replace('.', '-')
        else:
            from datetime import date
            display_date = date.today().strftime('%d-%m-%Y')
        
        fig.text(0.1, 0.82, f'Date:{display_date}', fontsize=10)
        if self.exam_type == 'Internal':
            fig.text(0.5, 0.82, f'Session: Morning', ha='center', fontsize=10)
        else:
            fig.text(0.5, 0.82, f'Session:{self.session}', ha='center', fontsize=10)
        fig.text(0.9, 0.82, f'Hall:{hall_no}', ha='right', fontsize=10)
        
        # Create column headers
        col_headers = [f'column {i+1}' for i in range(num_cols)]
        
        # Prepare table data based on exam type
        if self.exam_type in ['SEM', 'SEMESTER']:
            # Semester Exam: Simple grid with one student per cell
            table_data = [col_headers]
            for row_idx, row in enumerate(layout):
                table_data.append(row)
        else:
            # Internal Exam: Horizontal format (regno1 | regno2)
            table_data = [col_headers]
            for row_idx, row in enumerate(layout):
                row_data = []
                for cell in row:
                    if isinstance(cell, dict):
                        # Format as "regno1 | regno2" horizontally on same line
                        left = cell.get('left', '-')
                        right = cell.get('right', '-')
                        cell_text = f"{left} | {right}"
                        row_data.append(cell_text)
                    else:
                        row_data.append(str(cell))
                table_data.append(row_data)
        
        # Create main seating table
        table = ax.table(cellText=table_data, cellLoc='center', loc='center',
                        bbox=[0.1, 0.20, 0.8, 0.57])
        
        # Style the table
        table.auto_set_font_size(False)
        if self.exam_type == 'Internal':
            table.set_fontsize(7)  # Font size for readability
            table.scale(1, 1.8)  # Adjusted height for single-line horizontal format
        else:
            table.set_fontsize(9)
            table.scale(1, 2)
        
        # Style all cells with borders only (no colors)
        for key, cell in table.get_celld().items():
            cell.set_edgecolor('black')
            cell.set_linewidth(1)
            cell.set_facecolor('white')
            
            # Make header row bold
            if key[0] == 0:
                cell.set_text_props(weight='bold')
            
            # Style empty cells
            cell_text = cell.get_text().get_text()
            if "-" in cell_text:
                cell.set_text_props(color='black')
        
        # Add department breakdown table at bottom
        dept_data = [[dept, count] for dept, count in dept_counts.items()]
        dept_data.insert(0, ['Department', 'Count'])
        dept_data.append(['Total Number of Students:', str(occupied)])
        
        dept_table = ax.table(cellText=dept_data, cellLoc='left', loc='lower center',
                             bbox=[0.1, 0.05, 0.5, 0.15])
        dept_table.auto_set_font_size(False)
        dept_table.set_fontsize(9)
        dept_table.scale(1, 1.5)
        
        # Style department table
        for key, cell in dept_table.get_celld().items():
            cell.set_edgecolor('black')
            cell.set_linewidth(1)
            cell.set_facecolor('white')
            if key[0] == 0 or key[0] == len(dept_data) - 1:
                cell.set_text_props(weight='bold')
        
        plt.tight_layout()
        
        if save_path:
            plt.savefig(save_path, dpi=150, bbox_inches='tight', facecolor='white')
            plt.close()
        else:
            return fig
    
    def generate_student_pdf(self, output_file=None):
        """Generate student PDF with hall layouts (skip empty halls)"""
        if output_file is None:
            if self.exam_type == 'Internal':
                output_file = f'seating_student_{self.generation_date}_Y{self.year}_I{self.internal_number}.pdf'
            else:
                output_file = f'seating_student_{self.generation_date}_Y{self.year}_SEM_{self.session}.pdf'
        
        print("\n" + "=" * 60)
        print("GENERATING STUDENT PDF")
        print("=" * 60)
        
        # Filter out empty halls (halls with no students)
        non_empty_halls = [hall_no for hall_no in sorted(self.hall_wise_allocations.keys())
                          if len(self.hall_wise_allocations[hall_no]) > 0]
        
        if not non_empty_halls:
            print("Warning: No halls with students to generate PDF!")
            return None
        
        print(f"Generating PDF for {len(non_empty_halls)} halls with students...")
        
        with PdfPages(output_file) as pdf:
            for hall_no in non_empty_halls:
                print(f"  Creating layout for Hall {hall_no}...")
                fig = self.generate_hall_visual(hall_no)
                pdf.savefig(fig, bbox_inches='tight')
                plt.close(fig)
        
        print(f"\nStudent PDF generated: {output_file}")
        return output_file
    
    def generate_faculty_pdf(self, output_file=None):
        """Generate faculty PDF with summary table"""
        if output_file is None:
            if self.exam_type == 'Internal':
                output_file = f'seating_faculty_{self.generation_date}_Year{self.year}_Internal{self.internal_number}.pdf'
            else:
                output_file = f'seating_faculty_{self.generation_date}_Year{self.year}_SEM_{self.session}.pdf'
        
        print("\n" + "=" * 60)
        print("GENERATING FACULTY PDF")
        print("=" * 60)
        
        # Use portrait orientation for faculty summary
        doc = SimpleDocTemplate(output_file, pagesize=A4,
                               rightMargin=30, leftMargin=30,
                               topMargin=30, bottomMargin=30)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # College Header
        header_style = ParagraphStyle(
            'CollegeHeader',
            parent=styles['Heading1'],
            fontSize=16,
            fontName='Helvetica-Bold',
            alignment=TA_CENTER,
            spaceAfter=6
        )
        
        sub_header_style = ParagraphStyle(
            'SubHeader',
            parent=styles['Normal'],
            fontSize=11,
            alignment=TA_CENTER,
            spaceAfter=4
        )
        
        italic_style = ParagraphStyle(
            'Italic',
            parent=styles['Normal'],
            fontSize=9,
            fontName='Helvetica-Oblique',
            alignment=TA_CENTER,
            spaceAfter=6
        )
        
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=14,
            fontName='Helvetica-Bold',
            alignment=TA_CENTER,
            spaceAfter=12
        )
        
        elements.append(Paragraph("Marri Laxman Reddy Institute of Technology", header_style))
        elements.append(Paragraph("Hyderabad - 43", sub_header_style))
        elements.append(Paragraph("[An Autonomous Institution]", italic_style))
        elements.append(Paragraph("SEATING ARRANGEMENT - FACULTY SUMMARY", title_style))
        
        # Use exam_date for SEM exams, generation_date for Internal exams
        if self.exam_type in ['SEM', 'SEMESTER'] and self.exam_date:
            display_date = self.exam_date.replace('.', '-')
        else:
            display_date = self.generation_date
        
        elements.append(Paragraph(f"Date: {display_date} | Session: {self.session}", sub_header_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Overall Statistics
        total_students = len(self.allocations)
        halls_used = len(self.hall_wise_allocations)
        total_capacity = sum([self.halls_df[self.halls_df['hallno'] == h]['capacity'].values[0] 
                             for h in self.hall_wise_allocations.keys()])
        
        stats_data = [
            ['Overall Statistics', ''],
            ['Total Students', str(total_students)],
            ['Total Halls Used', f"{halls_used} / {len(self.halls_df)}"],
            ['Total Capacity', str(total_capacity)]
        ]
        
        stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
        stats_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(stats_table)
        elements.append(Spacer(1, 0.4*inch))
        
        # Hall Allocation Summary Table
        summary_title = Paragraph("Hall Allocation Summary", styles['Heading2'])
        elements.append(summary_title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Prepare table data - only include non-empty halls and prevent text overflow
        table_data = [['Hall', 'Cap', 'Occ', 'Invigilator', 'Departments']]
        
        # Filter to only non-empty halls
        non_empty_halls = [hall_no for hall_no in sorted(self.hall_wise_allocations.keys())
                          if len(self.hall_wise_allocations[hall_no]) > 0]
        
        for hall_no in non_empty_halls:
            hall_data = self.hall_wise_allocations[hall_no]
            capacity = self.halls_df[self.halls_df['hallno'] == hall_no]['capacity'].values[0]
            occupied = len(hall_data)
            
            # Get department counts - use compact format
            dept_counts = hall_data['Department'].value_counts()
            # Format as comma-separated to prevent overflow: "CSE:25,ECE:20,..."
            dept_breakdown = ', '.join([f"{dept}:{count}" for dept, count in dept_counts.items()])
            
            teacher = self.teacher_assignments.get(hall_no, "TBA")
            
            table_data.append([
                str(hall_no),
                str(capacity),
                str(occupied),
                teacher,
                dept_breakdown
            ])
        
        # Create table with adjusted widths to prevent overflow
        col_widths = [0.6*inch, 0.7*inch, 0.7*inch, 1.8*inch, 3.4*inch]  # Wider dept column
        summary_table = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        summary_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),  # Slightly smaller header
            ('FONTSIZE', (0, 1), (-1, -1), 7),  # Smaller data font
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('WORDWRAP', (0, 0), (-1, -1), True),  # Enable word wrap
        ]))
        
        elements.append(summary_table)
        
        # Build PDF
        doc.build(elements)
        
        print(f"\nFaculty PDF generated: {output_file}")
        return output_file
    
    def generate_excel_report(self, output_file='seating_allocation_report.xlsx'):
        """Generate comprehensive Excel report with multiple sheets"""
        
        print("\n" + "=" * 60)
        print("GENERATING EXCEL REPORT")
        print("=" * 60)
        
        writer = pd.ExcelWriter(output_file, engine='openpyxl')
        
        # Sheet 1: Complete allocation list (Linear Department)
        self.allocations.to_excel(writer, sheet_name='Complete Allocation', index=False)
        
        # Sheet 2: Hall-wise breakdown
        hall_summary = []
        for hall_no, hall_data in sorted(self.hall_wise_allocations.items()):
            dept_counts = hall_data['Department'].value_counts()
            hall_summary.append({
                'Hall No': hall_no,
                'Total Students': len(hall_data),
                'Departments': ', '.join([f"{dept}({count})" for dept, count in dept_counts.items()])
            })
        
        pd.DataFrame(hall_summary).to_excel(writer, sheet_name='Hall Summary', index=False)
        
        # Sheet 3: Department-wise summary
        dept_summary = self.allocations.groupby('Department').agg({
            'Register Number': 'count',
            'Hall No': lambda x: f"{x.min()} to {x.max()}"
        }).reset_index()
        dept_summary.columns = ['Department', 'Total Students', 'Hall Range']
        dept_summary.to_excel(writer, sheet_name='Department Summary', index=False)
        
        # Create individual hall sheets
        for hall_no, hall_data in sorted(self.hall_wise_allocations.items()):
            sheet_name = f"Hall {hall_no}"
            hall_data.to_excel(writer, sheet_name=sheet_name, index=False)
        
        writer.close()
        
        # Now format the Excel file
        self._format_excel(output_file)
        
        print(f"\nExcel report generated: {output_file}")
        print(f"Total sheets created: {3 + len(self.hall_wise_allocations)}")
        
        return output_file
    
    def _format_excel(self, file_path):
        """Apply formatting to the Excel file"""
        from openpyxl import load_workbook
        from openpyxl.styles import numbers
        
        wb = load_workbook(file_path)
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Find the Register Number column
            reg_num_col = None
            for idx, cell in enumerate(ws[1], 1):
                if cell.value == 'Register Number':
                    reg_num_col = idx
                    break
            
            # Format headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Format data cells
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for idx, cell in enumerate(row, 1):
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Force Register Number column to be text format
                    if reg_num_col and idx == reg_num_col:
                        cell.number_format = '@'  # Text format
                        # Ensure the value is stored as string
                        if cell.value is not None:
                            cell.value = str(cell.value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)
    
    def print_statistics(self):
        """Print allocation statistics"""
        print("\n" + "=" * 60)
        print("ALLOCATION STATISTICS")
        print("=" * 60)
        
        print("\nDepartment-wise allocation:")
        dept_stats = self.allocations.groupby('Department').agg({
            'Register Number': 'count',
            'Hall No': ['min', 'max']
        })
        
        for dept in dept_stats.index:
            count = dept_stats.loc[dept, ('Register Number', 'count')]
            hall_min = dept_stats.loc[dept, ('Hall No', 'min')]
            hall_max = dept_stats.loc[dept, ('Hall No', 'max')]
            print(f"  {dept:8s}: {count:3d} students (Halls {hall_min} to {hall_max})")
        
        print("\nHall utilization:")
        for hall_no in sorted(self.allocations['Hall No'].unique()):
            hall_capacity = self.halls_df[self.halls_df['hallno'] == hall_no]['capacity'].values[0]
            allocated = len(self.allocations[self.allocations['Hall No'] == hall_no])
            utilization = (allocated / hall_capacity) * 100
            print(f"  {hall_no}: {allocated:2d}/{hall_capacity:2d} seats ({utilization:5.1f}% utilized)")

    def save_allocation_to_db(self, cycle_id=None):
        """Save seating allocation to database for review
        
        Args:
            cycle_id: Exam cycle ID (optional, fetched from schedules if not provided)
            
        Returns:
            int: Number of records saved
        """
        if not hasattr(self, 'allocations') or self.allocations.empty:
            print("❌ No allocation to save. Run allocation first.")
            return 0
        
        if not self.exam_date or not self.session:
            print("❌ Cannot save allocation: exam_date and session are required for SEMESTER exams")
            return 0
        
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        try:
            # Get cycle_id from schedules if not provided
            if cycle_id is None and self.exam_date:
                cursor.execute('''
                    SELECT cycle_id 
                    FROM schedules 
                    WHERE exam_date = ? AND session = ?
                    LIMIT 1
                ''', (self.exam_date, self.session))
                result = cursor.fetchone()
                if result:
                    cycle_id = result[0]
            
            # Delete existing allocations for this date+session (if re-running)
            cursor.execute('''
                DELETE FROM seating_allocations 
                WHERE exam_date = ? AND session = ?
            ''', (self.exam_date, self.session))
            
            # Prepare data for insertion
            records_saved = 0
            for _, row in self.allocations.iterrows():
                # Get hall_id from hall name
                cursor.execute('SELECT hall_id FROM halls WHERE hall_name = ?', (row['Hall No'],))
                hall_result = cursor.fetchone()
                if not hall_result:
                    print(f"⚠️ Warning: Hall {row['Hall No']} not found in database")
                    continue
                hall_id = hall_result[0]
                
                # Get student_id from register number
                cursor.execute('SELECT student_id FROM students WHERE reg_no = ?', (row['Register Number'],))
                student_result = cursor.fetchone()
                if not student_result:
                    print(f"⚠️ Warning: Student {row['Register Number']} not found in database")
                    continue
                student_id = student_result[0]
                
                # Insert allocation record
                cursor.execute('''
                    INSERT INTO seating_allocations (
                        cycle_id, exam_date, session, hall_id, hall_name,
                        student_id, reg_no, student_name, department,
                        bench_number, seat_no, position, exam_type
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    cycle_id, self.exam_date, self.session, hall_id, row['Hall No'],
                    student_id, row['Register Number'], row['Name'], row['Department'],
                    row.get('Bench Number', 0), row['Seat No'], 
                    row.get('Position', 'N/A'), self.exam_type
                ))
                records_saved += 1
            
            conn.commit()
            print(f"\n✅ Saved {records_saved} seating allocations to database")
            print(f"   Exam Date: {self.exam_date} {self.session}")
            if cycle_id:
                print(f"   Cycle ID: {cycle_id}")
            
            return records_saved
            
        except Exception as e:
            conn.rollback()
            print(f"❌ Error saving allocation: {e}")
            import traceback
            traceback.print_exc()
            return 0
        finally:
            conn.close()
    
    @staticmethod
    def get_allocation_from_db(exam_date=None, session=None, cycle_id=None):
        """Retrieve seating allocation from database for review
        
        Args:
            exam_date: Exam date (DD.MM.YYYY)
            session: Session (FN/AN)
            cycle_id: Exam cycle ID (alternative to date+session)
            
        Returns:
            pd.DataFrame: Allocation data
        """
        conn = sqlite3.connect(DB_PATH)
        
        try:
            if exam_date and session:
                query = '''
                    SELECT 
                        allocation_id, cycle_id, exam_date, session,
                        hall_name, seat_no, reg_no, student_name, 
                        department, bench_number, exam_type, allocation_date
                    FROM seating_allocations
                    WHERE exam_date = ? AND session = ?
                    ORDER BY hall_name, bench_number, seat_no
                '''
                df = pd.read_sql_query(query, conn, params=(exam_date, session))
                
            elif cycle_id:
                query = '''
                    SELECT 
                        allocation_id, cycle_id, exam_date, session,
                        hall_name, seat_no, reg_no, student_name, 
                        department, bench_number, exam_type, allocation_date
                    FROM seating_allocations
                    WHERE cycle_id = ?
                    ORDER BY exam_date, session, hall_name, bench_number, seat_no
                '''
                df = pd.read_sql_query(query, conn, params=(cycle_id,))
            else:
                print("❌ Must provide either (exam_date + session) or cycle_id")
                return pd.DataFrame()
            
            if df.empty:
                print(f"⚠️ No allocation found for the specified criteria")
            else:
                print(f"✅ Retrieved {len(df)} seating records")
                if not df.empty:
                    print(f"   Exam: {df['exam_date'].iloc[0]} {df['session'].iloc[0]}")
                    print(f"   Halls: {df['hall_name'].nunique()}")
                    print(f"   Students: {df['reg_no'].nunique()}")
                    print(f"   Departments: {', '.join(df['department'].unique())}")
            
            return df
            
        except Exception as e:
            print(f"❌ Error retrieving allocation: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()
        finally:
            conn.close()
    
    @staticmethod
    def get_available_allocations():
        """Get list of all saved seating allocations
        
        Returns:
            pd.DataFrame: Summary of all allocations
        """
        conn = sqlite3.connect(DB_PATH)
        
        try:
            query = '''
                SELECT 
                    exam_date, session, cycle_id, exam_type,
                    COUNT(*) as student_count,
                    COUNT(DISTINCT hall_id) as hall_count,
                    COUNT(DISTINCT department) as dept_count,
                    MIN(allocation_date) as created_at
                FROM seating_allocations
                GROUP BY exam_date, session, cycle_id, exam_type
                ORDER BY exam_date, session
            '''
            df = pd.read_sql_query(query, conn)
            
            if df.empty:
                print("⚠️ No seating allocations found in database")
            else:
                print(f"\n📋 Found {len(df)} seating allocation(s):")
                for _, row in df.iterrows():
                    print(f"\n   {row['exam_date']} {row['session']} (Cycle: {row['cycle_id']})")
                    print(f"   └─ {row['student_count']} students, {row['hall_count']} halls, {row['dept_count']} departments")
            
            return df
            
        except Exception as e:
            print(f"❌ Error retrieving allocations: {e}")
            return pd.DataFrame()
        finally:
            conn.close()


def manage_faculty_selection(teachers_df):
    """Interactive faculty management"""
    print("\n" + "=" * 60)
    print("FACULTY/INVIGILATOR MANAGEMENT")
    print("=" * 60)
    
    selected_teachers = []
    available_teachers = teachers_df['Name'].str.strip().tolist()
    
    while True:
        print("\n" + "-" * 60)
        print("Options:")
        print("  [1] View all available faculty")
        print("  [2] Select faculty for invigilation")
        print("  [3] Add new faculty member")
        print("  [4] Remove selected faculty")
        print("  [5] Finish faculty selection")
        print("-" * 60)
        
        choice = input("\nEnter choice (1-5): ").strip()
        
        if choice == '1':
            print("\nAvailable Faculty:")
            for i, teacher in enumerate(available_teachers, 1):
                status = "[SELECTED]" if teacher in selected_teachers else "Available"
                print(f"  [{i}] {teacher:<30} ({status})")
        
        elif choice == '2':
            print("\nSelect Faculty (comma-separated numbers or 'all'):")
            for i, teacher in enumerate(available_teachers, 1):
                status = "[X]" if teacher in selected_teachers else "[ ]"
                print(f"  [{i}] [{status}] {teacher}")
            
            selection = input("\nEnter selection: ").strip().lower()
            if selection == 'all':
                selected_teachers = available_teachers.copy()
                print(f"Selected all {len(selected_teachers)} faculty members")
            else:
                try:
                    indices = [int(x.strip()) - 1 for x in selection.split(',')]
                    for idx in indices:
                        if 0 <= idx < len(available_teachers):
                            teacher = available_teachers[idx]
                            if teacher not in selected_teachers:
                                selected_teachers.append(teacher)
                    print(f"Total selected: {len(selected_teachers)} faculty")
                except:
                    print("Invalid input!")
        
        elif choice == '3':
            name = input("\nEnter new faculty name: ").strip()
            if name:
                available_teachers.append(name)
                selected_teachers.append(name)
                print(f"Added and selected: {name}")
        
        elif choice == '4':
            if not selected_teachers:
                print("No faculty selected yet!")
                continue
            print("\nSelected Faculty:")
            for i, teacher in enumerate(selected_teachers, 1):
                print(f"  [{i}] {teacher}")
            try:
                idx = int(input("\nEnter number to remove (0 to cancel): ").strip()) - 1
                if 0 <= idx < len(selected_teachers):
                    removed = selected_teachers.pop(idx)
                    print(f"Removed: {removed}")
            except:
                print("Invalid input!")
        
        elif choice == '5':
            if len(selected_teachers) == 0:
                print("Warning: No faculty selected!")
                confirm = input("Continue without faculty? (y/n): ").strip().lower()
                if confirm == 'y':
                    break
            else:
                print(f"\n{len(selected_teachers)} faculty members selected")
                break
        else:
            print("Invalid choice!")
    
    return selected_teachers
    
    # Get exam type from user
    print("\nExam Types:")
    print("  1. Internal - Continuous Internal Assessment (2 students per bench)")
    print("  2. SEM - End Semester Examination (1 student per bench)")
    exam_type_input = input("\nEnter exam type (Internal/SEM) [default: Internal]: ").strip().upper()
    if exam_type_input not in ['INTERNAL', 'SEM']:
        exam_type = 'Internal'
    else:
        exam_type = 'Internal' if exam_type_input == 'INTERNAL' else 'SEM'
    
    # Get internal exam number if Internal exam is selected
    internal_number = 1
    session = 'FN'  # Default session
    
    if exam_type == 'Internal':
        internal_input = input("\nWhich Internal Exam? (1/2) [default: 1]: ").strip()
        if internal_input in ['1', '2']:
            internal_number = int(internal_input)
        else:
            internal_number = 1
        print(f"Selected: Internal {internal_number} (Morning session)")
    else:
        # Get session only for SEM exams
        session = input("\nEnter session (FN/AN) [default: FN]: ").strip().upper()
        if session not in ['FN', 'AN']:
            session = 'FN'
    
    # Faculty/Invigilator Management
    selected_teachers = manage_faculty_selection(teachers_df)
    
    # Hall Selection
    selected_halls = manage_hall_selection(halls_df, total_students, exam_type)
    
    # Final Confirmation
    print("\n" + "=" * 60)
    print("CONFIGURATION SUMMARY")
    print("=" * 60)
    print(f"Year: {year_names[year]} Year")
    print(f"Exam Type: {exam_type} {internal_number if exam_type == 'Internal' else ''}")
    print(f"Session: {session}")
    print(f"Total Students: {total_students}")
    print(f"Selected Halls: {len(selected_halls)} halls")
    print(f"Selected Faculty: {len(selected_teachers)} invigilators")
    print("=" * 60)
    
    confirm = input("\nProceed with allocation generation? (y/n): ").strip().lower()
    if confirm != 'y':
        print("\n✗ Allocation cancelled")
        return
    
    # Create allocation system with selected halls and teachers
    system = SeatingAllocationSystem(halls_file, students_file, teachers_file, 
                                     session=session, exam_type=exam_type, year=year, 
                                     internal_number=internal_number,
                                     selected_halls=selected_halls,
                                     selected_teachers=selected_teachers)

def manage_hall_selection(halls_df, total_students, exam_type):
    """Interactive hall selection with dynamic capacity tracking"""
    print("\n" + "=" * 60)
    print("HALL SELECTION")
    print("=" * 60)
    print(f"\nTotal Students: {total_students}")
    
    # Calculate effective capacity per hall
    if exam_type == 'Internal':
        halls_df['effective_capacity'] = halls_df['capacity'] * 2
        print("Exam Type: Internal (2 students per bench)")
    else:
        halls_df['effective_capacity'] = halls_df['capacity']
        print("Exam Type: SEM (1 student per bench)")
    
    selected_halls = []
    accumulated_capacity = 0
    
    while True:
        remaining = total_students - accumulated_capacity
        print("\n" + "-" * 60)
        print(f"Current Status:")
        print(f"  Students to accommodate: {total_students}")
        print(f"  Current capacity: {accumulated_capacity}")
        print(f"  Remaining: {remaining}")
        print(f"  Halls selected: {len(selected_halls)}")
        print("-" * 60)
        
        print("\nOptions:")
        print("  [1] View all available halls")
        print("  [2] Select halls")
        print("  [3] Auto-select optimal halls")
        print("  [4] Remove selected hall")
        print("  [5] Finish hall selection")
        print("-" * 60)
        
        choice = input("\nEnter choice (1-5): ").strip()
        
        if choice == '1':
            print("\nAvailable Halls:")
            print(f"{'No':<6} {'Capacity':<10} {'Benches':<10} {'Students':<12} {'Status':<15}")
            print("-" * 60)
            for _, hall in halls_df.iterrows():
                hall_no = hall['hallno']
                cap = int(hall['capacity'])
                eff_cap = int(hall['effective_capacity'])
                status = "[SELECTED]" if hall_no in selected_halls else "Available"
                print(f"{hall_no:<6} {cap:<10} {cap:<10} {eff_cap:<12} {status:<15}")
        
        elif choice == '2':
            print("\nSelect Halls (comma-separated hall numbers or names):")
            hall_input = input("Hall numbers: ").strip()
            try:
                # Support both "1,2,3" and "Hall 1, Hall 2, Hall 3" formats
                hall_inputs = [x.strip() for x in hall_input.split(',')]
                for hall_input_item in hall_inputs:
                    # Try to match as integer (convert to "Hall X" format)
                    try:
                        hall_num = int(hall_input_item)
                        hall_name = f"Hall {hall_num}"
                    except ValueError:
                        # Use as-is if not a number
                        hall_name = hall_input_item
                    
                    if hall_name in halls_df['hallno'].values and hall_name not in selected_halls:
                        selected_halls.append(hall_name)
                        cap = int(halls_df[halls_df['hallno'] == hall_name]['effective_capacity'].values[0])
                        accumulated_capacity += cap
                        print(f"Added {hall_name} (capacity: {cap})")
                    elif hall_name in selected_halls:
                        print(f"WARNING: {hall_name} already selected")
                    else:
                        print(f"WARNING: {hall_name} not found")
                
                print(f"\nTotal capacity now: {accumulated_capacity}/{total_students}")
                if accumulated_capacity >= total_students:
                    print("Sufficient capacity achieved!")
            except Exception as e:
                print(f"Invalid input! Error: {e}")
        
        elif choice == '3':
            # Auto-select halls to minimize waste using optimized algorithm
            if selected_halls:
                print(f"\nWarning: You already have {len(selected_halls)} halls selected (capacity: {accumulated_capacity})")
                confirm = input("Replace with auto-selection? (y/n): ").strip().lower()
                if confirm != 'y':
                    print("Keeping existing selection.")
                    continue
            
            print("\nOptimizing hall selection...")
            
            # Sort halls by capacity (largest first for better packing)
            halls_sorted = halls_df.sort_values('effective_capacity', ascending=False).reset_index(drop=True)
            
            best_selection = None
            best_waste = float('inf')
            best_count = float('inf')
            
            # Try greedy approach with largest halls first
            selection1 = []
            capacity1 = 0
            for _, hall in halls_sorted.iterrows():
                if capacity1 >= total_students:
                    break
                selection1.append(hall['hallno'])
                capacity1 += int(hall['effective_capacity'])
            
            waste1 = capacity1 - total_students
            if waste1 >= 0 and (len(selection1) < best_count or (len(selection1) == best_count and waste1 < best_waste)):
                best_selection = selection1
                best_waste = waste1
                best_count = len(selection1)
            
            # Try greedy approach with smallest halls first
            halls_sorted_asc = halls_df.sort_values('effective_capacity', ascending=True).reset_index(drop=True)
            selection2 = []
            capacity2 = 0
            for _, hall in halls_sorted_asc.iterrows():
                if capacity2 >= total_students:
                    break
                selection2.append(hall['hallno'])
                capacity2 += int(hall['effective_capacity'])
            
            waste2 = capacity2 - total_students
            if waste2 >= 0 and (len(selection2) < best_count or (len(selection2) == best_count and waste2 < best_waste)):
                best_selection = selection2
                best_waste = waste2
                best_count = len(selection2)
            
            # Try to find better combination by removing and replacing halls
            if best_selection:
                temp_selection = best_selection.copy()
                temp_capacity = sum(int(halls_df[halls_df['hallno'] == h]['effective_capacity'].values[0]) for h in temp_selection)
                
                # Try removing largest hall and adding smaller ones if it reduces waste
                for hall_to_remove in temp_selection:
                    hall_cap = int(halls_df[halls_df['hallno'] == hall_to_remove]['effective_capacity'].values[0])
                    new_capacity = temp_capacity - hall_cap
                    
                    if new_capacity >= total_students:
                        new_selection = [h for h in temp_selection if h != hall_to_remove]
                        new_waste = new_capacity - total_students
                        if len(new_selection) < best_count or (len(new_selection) == best_count and new_waste < best_waste):
                            best_selection = new_selection
                            best_waste = new_waste
                            best_count = len(new_selection)
                            temp_selection = new_selection
                            temp_capacity = new_capacity
            
            # Update the actual selections
            if best_selection:
                selected_halls = best_selection
                accumulated_capacity = sum(int(halls_df[halls_df['hallno'] == h]['effective_capacity'].values[0]) for h in selected_halls)
                
                print(f"Optimized selection: {len(selected_halls)} halls")
                print(f"  Total capacity: {accumulated_capacity}")
                print(f"  Waste: {accumulated_capacity - total_students} seats ({((accumulated_capacity - total_students) / total_students * 100):.1f}%)")
                print(f"  Halls: {', '.join(map(str, selected_halls))}")
            else:
                print("ERROR: Could not find optimal selection")
        
        elif choice == '4':
            if not selected_halls:
                print("No halls selected yet!")
                continue
            print("\nSelected Halls:")
            for i, hall_name in enumerate(selected_halls, 1):
                cap = int(halls_df[halls_df['hallno'] == hall_name]['effective_capacity'].values[0])
                print(f"  [{i}] {hall_name} (capacity: {cap})")
            try:
                idx = int(input("\nEnter number to remove (0 to cancel): ").strip()) - 1
                if 0 <= idx < len(selected_halls):
                    removed_hall = selected_halls.pop(idx)
                    cap = int(halls_df[halls_df['hallno'] == removed_hall]['effective_capacity'].values[0])
                    accumulated_capacity -= cap
                    print(f"Removed {removed_hall}")
                    print(f"  New capacity: {accumulated_capacity}")
            except:
                print("Invalid input!")
        
        elif choice == '5':
            if accumulated_capacity < total_students:
                print(f"\n⚠ Warning: Insufficient capacity!")
                print(f"   Need: {total_students}, Have: {accumulated_capacity}")
                confirm = input("Continue anyway? (y/n): ").strip().lower()
                if confirm != 'y':
                    continue
            print(f"\n{len(selected_halls)} halls selected (capacity: {accumulated_capacity})")
            break
        else:
            print("Invalid choice!")
    
    return selected_halls


def create_internal_schedule(year, semester_type, internal_number):
    """Create internal exam schedule with date-based slots (not department-based)"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Create exam cycle
    current_date = datetime.now().strftime('%Y-%m-%d')
    cursor.execute('''
        INSERT INTO exam_cycles (exam_type, year_group, start_date, end_date, created_date, status)
        VALUES (?, ?, ?, ?, ?, ?)
    ''', ('INTERNAL', year, current_date, current_date, current_date, 'ACTIVE'))
    cycle_id = cursor.lastrowid
    
    # Get all subjects for this year and semester
    cursor.execute('''
        SELECT subject_id, subject_code, subject_name, department
        FROM subjects
        WHERE year = ? AND semester_type = ? AND exam_type IN ('BOTH', 'INTERNAL')
        ORDER BY subject_code
    ''', (year, semester_type))
    subjects = cursor.fetchall()
    
    if not subjects:
        print(f"\n❌ No subjects found for Year {year} {semester_type} semester")
        conn.close()
        return None
    
    # Ask for number of days/slots
    print(f"\nTotal subjects to schedule: {len(subjects)}")
    print("How many days/slots do you want to spread the exams across?")
    while True:
        try:
            num_slots = int(input("Enter number of slots (1-10): ").strip())
            if 1 <= num_slots <= 10:
                break
            else:
                print("Please enter a number between 1 and 10")
        except ValueError:
            print("Please enter a valid number")
    
    # Distribute subjects evenly across slots
    subjects_per_slot = len(subjects) // num_slots
    remainder = len(subjects) % num_slots
    
    slot_dates = []
    base_date = datetime.now()
    for i in range(num_slots):
        # Generate dates (skip Sundays)
        exam_date = base_date
        while exam_date.weekday() == 6:  # Skip Sundays
            exam_date += timedelta(days=1)
        slot_dates.append(exam_date.strftime('%d.%m.%Y'))
        base_date = exam_date + timedelta(days=1)
    
    # Schedule subjects to dates
    subject_index = 0
    schedule_summary = {}
    
    for slot_num, slot_date in enumerate(slot_dates, 1):
        # Calculate how many subjects for this slot
        count = subjects_per_slot + (1 if slot_num <= remainder else 0)
        slot_subjects = subjects[subject_index:subject_index + count]
        subject_index += count
        
        session_name = f"SLOT{slot_num}"
        schedule_summary[session_name] = {
            'date': slot_date,
            'subjects': len(slot_subjects)
        }
        
        # Insert into schedules
        for subject in slot_subjects:
            cursor.execute('''
                INSERT INTO schedules (cycle_id, subject_id, exam_date, session)
                VALUES (?, ?, ?, ?)
            ''', (cycle_id, subject[0], slot_date, session_name))
    
    conn.commit()
    conn.close()
    
    print(f"\n✓ Internal {internal_number} schedule created:")
    print(f"  • Semester: {semester_type}")
    print(f"  • Total slots: {num_slots}")
    for slot, info in schedule_summary.items():
        print(f"    - {slot} ({info['date']}): {info['subjects']} subjects")
    
    return cycle_id


def main_internal_exam():
    """Handle Internal exam allocation (read from existing schedule)"""
    print("\n" + "=" * 60)
    print("INTERNAL EXAM - SEATING ALLOCATION")
    print("=" * 60)
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Step 1: Get available internal exam cycles
    cursor.execute('''
        SELECT DISTINCT ec.cycle_id, ec.year_group, ec.start_date, ec.end_date, 
               ec.created_date, ec.status,
               COUNT(DISTINCT sch.schedule_id) as schedule_count
        FROM exam_cycles ec
        LEFT JOIN schedules sch ON ec.cycle_id = sch.cycle_id
        WHERE ec.exam_type = 'INTERNAL'
        GROUP BY ec.cycle_id
        ORDER BY ec.created_date DESC
    ''')
    cycles = cursor.fetchall()
    
    if not cycles:
        print("\n❌ No internal exam cycles found!")
        print("Please create a schedule first using main.py")
        conn.close()
        return
    
    print("\nAvailable Internal Exam Cycles:")
    for idx, (cycle_id, year, start_date, end_date, created, status, sched_count) in enumerate(cycles, 1):
        year_names = {1: "First", 2: "Second", 3: "Third", 4: "Fourth"}
        print(f"{idx}. Cycle {cycle_id} - {year_names.get(year, f'Year {year}')} | {start_date} to {end_date} | {sched_count} schedules | Status: {status}")
    
    cycle_input = input(f"\nSelect cycle (1-{len(cycles)}): ").strip()
    
    if not cycle_input.isdigit() or int(cycle_input) < 1 or int(cycle_input) > len(cycles):
        print("\n❌ Invalid cycle selection")
        conn.close()
        return
    
    selected_cycle = cycles[int(cycle_input) - 1]
    cycle_id = selected_cycle[0]
    year = selected_cycle[1]
    
    year_names = {1: "First", 2: "Second", 3: "Third", 4: "Fourth"}
    print(f"\n✓ {year_names.get(year, f'Year {year}')} Year cycle selected")
    
    # Step 2: Get available slots (dates/sessions) from this cycle
    cursor.execute('''
        SELECT DISTINCT sch.exam_date, sch.session,
               COUNT(DISTINCT sub.subject_id) as subject_count,
               COUNT(DISTINCT s.student_id) as student_count
        FROM schedules sch
        JOIN subjects sub ON sch.subject_id = sub.subject_id
        LEFT JOIN student_subjects ss ON sub.subject_id = ss.subject_id
        LEFT JOIN students s ON ss.student_id = s.student_id AND s.active = 1
        WHERE sch.cycle_id = ?
        GROUP BY sch.exam_date, sch.session
        ORDER BY sch.exam_date, sch.session
    ''', (cycle_id,))
    slots = cursor.fetchall()
    
    if not slots:
        print("\n❌ No scheduled exams found in this cycle")
        conn.close()
        return
    
    print("\n" + "=" * 60)
    print("SELECT SLOT FOR SEATING ALLOCATION")
    print("=" * 60)
    print("\nAvailable Slots:")
    for idx, (exam_date, session, subj_count, stud_count) in enumerate(slots, 1):
        print(f"{idx}. {exam_date} {session} - {subj_count} subjects, {stud_count} students")
    
    slot_input = input(f"\nSelect slot (1-{len(slots)}): ").strip()
    
    if not slot_input.isdigit() or int(slot_input) < 1 or int(slot_input) > len(slots):
        print("\n❌ Invalid slot selection")
        conn.close()
        return
    
    selected_slot = slots[int(slot_input) - 1]
    exam_date = selected_slot[0]
    session = selected_slot[1]
    slot_students = selected_slot[3]
    
    print(f"\n✓ Slot selected: {exam_date} {session} ({slot_students} students)")
    
    # Get students for this slot
    cursor.execute('''
        SELECT DISTINCT s.student_id, s.reg_no, s.name, s.department
        FROM students s
        JOIN student_subjects ss ON s.student_id = ss.student_id
        JOIN schedules sch ON ss.subject_id = sch.subject_id
        WHERE sch.cycle_id = ? AND sch.exam_date = ? AND sch.session = ? AND s.active = 1
        ORDER BY s.department, s.reg_no
    ''', (cycle_id, exam_date, session))
    students = cursor.fetchall()
    total_students = len(students)
    
    # Load halls and teachers
    halls_df = pd.read_sql_query("SELECT hall_name as hallno, capacity, columns FROM halls WHERE active = 1", conn)
    teachers_df = pd.read_sql_query("SELECT teacher_name as Name, department as Department FROM teachers WHERE active = 1", conn)
    conn.close()
    
    # Hall Selection
    print("\n" + "=" * 60)
    print("HALL SELECTION")
    print("=" * 60)
    selected_halls = manage_hall_selection(halls_df, total_students, 'Internal')
    
    # Faculty Management
    print("\n" + "=" * 60)
    print("FACULTY MANAGEMENT")
    print("=" * 60)
    print(f"Halls selected: {len(selected_halls)}")
    print(f"Minimum faculty required: {len(selected_halls)}")
    selected_teachers = manage_faculty_selection(teachers_df)
    
    # Create allocation system
    system = SeatingAllocationSystem(
        use_database=True,
        session=session,
        exam_type='Internal',
        year=year,
        internal_number=1,  # Not relevant when using existing schedule
        selected_halls=selected_halls,
        selected_teachers=selected_teachers,
        exam_date=exam_date
    )
    
    # Final confirmation
    print("\n" + "=" * 60)
    print("FINAL CONFIRMATION")
    print("=" * 60)
    print(f"Year: {year_names.get(year, f'Year {year}')}")
    print(f"Exam Date: {exam_date}")
    print(f"Session: {session}")
    print(f"Students: {total_students}")
    print(f"Mode: 2 students per bench")
    print(f"Selected Halls: {len(selected_halls)}")
    print(f"Selected Teachers: {len(selected_teachers)}")
    
    confirm = input("\nProceed with allocation and PDF generation? (yes/no) [default: yes]: ").strip().lower()
    if confirm not in ['yes', 'y', '']:
        print("\n✗ Allocation cancelled")
        return
    
    # Generate allocation
    print(f"\nGenerating seating arrangement for {exam_date} {session}...")
    allocations = system.allocate_seats_mixed_department()
    system.assign_teachers()
    
    # Save to database
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    records_saved = 0
    for _, alloc in system.allocations.iterrows():
        try:
            # Get student_id from reg_no
            cursor.execute('SELECT student_id FROM students WHERE reg_no = ?', (alloc.get('Register Number'),))
            student_result = cursor.fetchone()
            if not student_result:
                continue
            student_id = student_result[0]
            
            # Get hall_id from hall_name
            cursor.execute('SELECT hall_id FROM halls WHERE hall_name = ?', (alloc.get('Hall No') or alloc.get('Hall'),))
            hall_result = cursor.fetchone()
            hall_id = hall_result[0] if hall_result else None
            
            cursor.execute('''
                INSERT INTO seating_allocations (
                    cycle_id, exam_date, session, hall_id, hall_name, 
                    student_id, reg_no, student_name, department,
                    bench_number, seat_no, position, exam_type, allocation_date
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                cycle_id,
                exam_date,
                session,
                hall_id,
                alloc.get('Hall No') or alloc.get('Hall'),
                student_id,
                alloc.get('Register Number'),
                alloc.get('Name'),
                alloc.get('Department'),
                alloc.get('Bench No') or alloc.get('Bench'),
                str(alloc.get('Bench No') or alloc.get('Bench')) + alloc.get('Position', 'A'),
                alloc.get('Position', 'A'),
                'Internal',
                datetime.now().strftime('%Y-%m-%d')
            ))
            records_saved += 1
        except (sqlite3.IntegrityError, Exception) as e:
            print(f"Warning: Could not save allocation: {e}")
            pass
    
    conn.commit()
    conn.close()
    
    # Generate PDFs
    student_pdf = system.generate_student_pdf()
    faculty_pdf = system.generate_faculty_pdf()
    system.print_statistics()
    
    print("\n" + "=" * 60)
    print("ALLOCATION COMPLETE!")
    print("=" * 60)
    print(f"\n✓ Saved {records_saved} seat allocations to database")
    print(f"\nGenerated Files:")
    print(f"  • Student PDF: {student_pdf}")
    print(f"  • Faculty PDF: {faculty_pdf}")
    print("\n")


def main_sem_exam():
    """Handle SEM exam allocation (new hierarchical workflow)"""
    print("\n" + "=" * 60)
    print("SEMESTER EXAM - SEATING ALLOCATION")
    print("=" * 60)
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Step 1: Get Year and Semester
    print("\nStep 1: Select Year and Semester")
    print("=" * 40)
    print("Years:")
    print("  1. First Year")
    print("  2. Second Year")
    print("  3. Third Year")
    print("  4. Fourth Year")
    year_input = input("\nEnter year (1/2/3/4): ").strip()
    
    if year_input not in ['1', '2', '3', '4']:
        print("❌ Invalid year selection")
        conn.close()
        return
    
    year = int(year_input)
    year_names = {1: "First", 2: "Second", 3: "Third", 4: "Fourth"}
    
    print("\nSemesters:")
    print("  1. ODD Semester")
    print("  2. EVEN Semester")
    sem_input = input("\nEnter semester (1/2): ").strip()
    
    if sem_input not in ['1', '2']:
        print("❌ Invalid semester selection")
        conn.close()
        return
    
    semester = 'ODD' if sem_input == '1' else 'EVEN'
    
    print(f"\n✓ Selected: {year_names[year]} Year, {semester} Semester")
    
    # Step 2: Get available dates and sessions from schedules
    print("\n" + "=" * 60)
    print("Step 2: Select Date and Session")
    print("=" * 60)
    
    # Query schedules for this year (all semesters - students can have arrears from both)
    cursor.execute('''
        SELECT DISTINCT sch.exam_date, sch.session
        FROM schedules sch
        JOIN subjects sub ON sch.subject_id = sub.subject_id
        WHERE sub.year = ?
        ORDER BY sch.exam_date, sch.session
    ''', (year,))
    
    available_slots = cursor.fetchall()
    
    if not available_slots:
        print(f"\n❌ No exam schedule found for Year {year}, {semester} semester")
        print(f"\nPlease run the scheduler first to create exam schedule.")
        conn.close()
        return
    
    print(f"\nAvailable Exam Dates and Sessions:")
    print("-" * 40)
    for idx, (exam_date, session) in enumerate(available_slots, 1):
        # Get subject count for this slot (all semesters)
        cursor.execute('''
            SELECT COUNT(DISTINCT sub.subject_id)
            FROM schedules sch
            JOIN subjects sub ON sch.subject_id = sub.subject_id
            WHERE sub.year = ?
                AND sch.exam_date = ? AND sch.session = ?
        ''', (year, exam_date, session))
        subject_count = cursor.fetchone()[0]
        
        print(f"  {idx}. {exam_date} - {session} ({subject_count} subjects)")
    
    slot_input = input(f"\nSelect slot (1-{len(available_slots)}): ").strip()
    
    try:
        slot_idx = int(slot_input) - 1
        if slot_idx < 0 or slot_idx >= len(available_slots):
            raise ValueError
        
        exam_date, session = available_slots[slot_idx]
    except:
        print("❌ Invalid slot selection")
        conn.close()
        return
    
    print(f"\n✓ Selected: {exam_date} - {session}")
    
    # Step 3: Show subjects scheduled for this date+session
    print("\n" + "=" * 60)
    print("Step 3: Subjects Scheduled")
    print("=" * 60)
    
    cursor.execute('''
        SELECT DISTINCT sub.subject_code, sub.subject_name
        FROM schedules sch
        JOIN subjects sub ON sch.subject_id = sub.subject_id
        WHERE sub.year = ?
            AND sch.exam_date = ? AND sch.session = ?
        ORDER BY sub.subject_code
    ''', (year, exam_date, session))
    
    scheduled_subjects = cursor.fetchall()
    
    print("\nSubjects on this date and session:")
    for subject_code, subject_name in scheduled_subjects:
        print(f"  • {subject_code} - {subject_name}")
    
    # Step 4: Collect students (regular + arrear)
    print("\n" + "=" * 60)
    print("Step 4: Collecting Students")
    print("=" * 60)
    
    # Get subject IDs for scheduled subjects
    subject_codes = [sc[0] for sc in scheduled_subjects]
    placeholders = ','.join(['?' for _ in subject_codes])
    
    # Query: Regular students (enrolled in these subjects)
    cursor.execute(f'''
        SELECT DISTINCT s.student_id, s.reg_no, s.name, s.department, s.year
        FROM students s
        JOIN student_subjects ss ON s.student_id = ss.student_id
        JOIN subjects sub ON ss.subject_id = sub.subject_id
        WHERE sub.subject_code IN ({placeholders})
            AND s.active = 1
            AND ss.is_arrear = 0
        ORDER BY s.department, s.reg_no
    ''', subject_codes)
    
    regular_students = cursor.fetchall()
    
    # Query: Arrear students (have arrears in these subjects)
    cursor.execute(f'''
        SELECT DISTINCT s.student_id, s.reg_no, s.name, s.department, s.year, s.arrears
        FROM students s
        WHERE s.active = 1 AND s.arrears IS NOT NULL
    ''')
    
    all_students_with_arrears = cursor.fetchall()
    arrear_students = []
    
    for student_id, reg_no, name, dept, student_year, arrears_json in all_students_with_arrears:
        if arrears_json:
            try:
                import json
                arrears_list = json.loads(arrears_json)
                # Check if any scheduled subject is in arrears
                if any(sc in arrears_list for sc in subject_codes):
                    arrear_students.append((student_id, reg_no, name, dept, student_year))
            except:
                pass
    
    # Combine regular + arrear students
    all_student_data = list(regular_students) + arrear_students
    # Remove duplicates by student_id
    unique_students = {}
    for student_id, reg_no, name, dept, student_year in all_student_data:
        if student_id not in unique_students:
            unique_students[student_id] = (reg_no, name, dept, student_year)
    
    total_students = len(unique_students)
    
    print(f"\n✓ Regular students: {len(regular_students)}")
    print(f"✓ Arrear students: {len(arrear_students)}")
    print(f"✓ Total students: {total_students}")
    
    if total_students == 0:
        print("\n❌ No students found for this exam slot")
        conn.close()
        return
    
    # Step 5: Hall and Faculty Selection
    halls_df = pd.read_sql_query("SELECT hall_name as hallno, capacity, columns FROM halls WHERE active = 1", conn)
    teachers_df = pd.read_sql_query("SELECT teacher_name as Name, department as Department FROM teachers WHERE active = 1", conn)
    
    print("\n" + "=" * 60)
    print("Step 5: Hall Selection")
    print("=" * 60)
    selected_halls = manage_hall_selection(halls_df, total_students, 'SEMESTER')
    
    print("\n" + "=" * 60)
    print("Step 6: Faculty Management")
    print("=" * 60)
    print(f"Halls selected: {len(selected_halls)}")
    print(f"Minimum faculty required: {len(selected_halls)}")
    selected_teachers = manage_faculty_selection(teachers_df)
    
    conn.close()
    
    # Step 6: Create allocation system with exam_date
    print("\n" + "=" * 60)
    print("FINAL CONFIRMATION")
    print("=" * 60)
    print(f"Year: {year_names[year]} Year")
    print(f"Semester: {semester}")
    print(f"Exam Date: {exam_date}")
    print(f"Session: {session}")
    print(f"Subjects: {len(scheduled_subjects)}")
    print(f"Total Students: {total_students}")
    print(f"Mode: 1 student per bench")
    print(f"Selected Halls: {len(selected_halls)}")
    print(f"Selected Teachers: {len(selected_teachers)}")
    
    confirm = input("\nProceed with allocation and PDF generation? (yes/no) [default: yes]: ").strip().lower()
    if confirm not in ['yes', 'y', '']:
        print("\n✗ Allocation cancelled")
        return
    
    # Create allocation system
    system = SeatingAllocationSystem(
        use_database=True,
        session=session,
        exam_type='SEMESTER',
        year=year,
        exam_date=exam_date,
        selected_halls=selected_halls,
        selected_teachers=selected_teachers
    )
    
    # Generate allocation
    print(f"\nGenerating seating arrangement for {exam_date} {session}...")
    allocations = system.allocate_seats_mixed_department()
    system.assign_teachers()
    
    # Save to database
    print("\nSaving allocation to database...")
    records_saved = system.save_allocation_to_db()
    
    # Generate PDFs
    student_pdf = system.generate_student_pdf()
    faculty_pdf = system.generate_faculty_pdf()
    system.print_statistics()
    
    print("\n" + "=" * 60)
    print("ALLOCATION COMPLETE!")
    print("=" * 60)
    print(f"\n✓ Saved {records_saved} seat allocations to database")
    print(f"\nGenerated Files:")
    print(f"  • Student PDF: {student_pdf}")
    print(f"  • Faculty PDF: {faculty_pdf}")
    print("\nReview allocation: python review_seating_allocations.py")
    print("\n")


def main():
    """Main execution function"""
    print("\n" + "=" * 60)
    print("SEATING ARRANGEMENT ALLOCATION SYSTEM")
    print("Academic Year 2024-25")
    print("=" * 60)
    
    # Check if database exists
    if not os.path.exists(DB_PATH):
        print(f"\n❌ ERROR: Database not found!")
        print(f"Expected location: {DB_PATH}")
        print(f"\nPlease run: python integrated_db_setup.py")
        print(f"from the 'Exam Scheduling Algorithm' folder first.")
        return
    
    # Step 1: Ask exam type
    print("\n" + "=" * 60)
    print("SELECT EXAM TYPE")
    print("=" * 60)
    print("\n1. Internal Exam")
    print("   • Continuous Internal Assessment")
    print("   • 2 students per bench")
    print("   • Fixed year-based allocation")
    print("\n2. SEM Exam")
    print("   • End Semester Examination")
    print("   • 1 student per bench")
    print("   • Schedule-based allocation (includes arrear students)")
    
    exam_type_input = input("\nSelect exam type (1/2): ").strip()
    
    if exam_type_input == '1':
        main_internal_exam()
    elif exam_type_input == '2':
        main_sem_exam()
    else:
        print("\n❌ Invalid selection")
        return


if __name__ == "__main__":
    main()
