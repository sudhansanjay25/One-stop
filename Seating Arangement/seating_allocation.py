"""
Seating Arrangement Allocation System
--------------------------------------
This system allocates students to exam halls with the following features:
1. Linear department seating allocation
2. Teacher/Invigilator assignment
3. PDF generation (Student & Faculty versions)
4. Excel output with comprehensive reports
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.backends.backend_pdf import PdfPages
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT


class SeatingAllocationSystem:
    def __init__(self, halls_file, students_file, teachers_file, session='FN'):
        """Initialize the seating allocation system"""
        # Read halls data with columns information
        self.halls_df = pd.read_csv(halls_file)
        self.halls_df.columns = self.halls_df.columns.str.strip()
        
        # Read students data - preserve register numbers as strings
        self.students_df = pd.read_csv(students_file, dtype={'Register Number': str})
        self.students_df.columns = self.students_df.columns.str.strip()
        
        # Read teachers data
        self.teachers_df = pd.read_csv(teachers_file)
        self.teachers_df.columns = self.teachers_df.columns.str.strip()
        
        # Prepare data structures
        self.allocations = []
        self.hall_wise_allocations = {}
        self.teacher_assignments = {}
        self.session = session  # 'FN' or 'AN'
        self.generation_date = datetime.now().strftime('%Y-%m-%d')
        
    def allocate_seats_mixed_department(self):
        """
        Allocate seats linearly by department
        Complete one department before moving to the next
        """
        print("=" * 60)
        print("SEATING ALLOCATION - LINEAR DEPARTMENT FORMAT")
        print("=" * 60)
        
        # Sort students by department and register number
        students_sorted = self.students_df.sort_values(
            ['Department', 'Register Number']
        ).reset_index(drop=True)
        
        # Allocate students to halls
        current_hall_idx = 0
        current_seat_in_hall = 1
        
        allocations = []
        
        for idx, student in students_sorted.iterrows():
            hall_no = self.halls_df.loc[current_hall_idx, 'hallno']
            hall_capacity = self.halls_df.loc[current_hall_idx, 'capacity']
            
            allocations.append({
                'Hall No': hall_no,
                'Seat No': current_seat_in_hall,
                'Register Number': student['Register Number'],
                'Name': student['Name'],
                'Department': student['Department']
            })
            
            current_seat_in_hall += 1
            
            # Check if we need to move to next hall
            if current_seat_in_hall > hall_capacity:
                current_hall_idx += 1
                current_seat_in_hall = 1
                
                if current_hall_idx >= len(self.halls_df):
                    print("Warning: Ran out of halls!")
                    break
        
        self.allocations = pd.DataFrame(allocations)
        print(f"\nTotal students allocated: {len(self.allocations)}")
        print(f"Halls used: {current_hall_idx + 1} out of {len(self.halls_df)}")
        
        # Create hall-wise summary
        self._create_hall_wise_summary()
        
        return self.allocations
    
    def allocate_seats_alternating_department(self):
        """
        Allocate seats with different departments alternating
        This ensures students from the same department don't sit next to each other
        """
        print("\n" + "=" * 60)
        print("SEATING ALLOCATION - ALTERNATING DEPARTMENT FORMAT")
        print("=" * 60)
        
        # Group students by department
        departments = self.students_df['Department'].unique()
        dept_groups = {dept: self.students_df[self.students_df['Department'] == dept].copy() 
                      for dept in departments}
        
        # Sort each department group by register number
        for dept in dept_groups:
            dept_groups[dept] = dept_groups[dept].sort_values('Register Number').reset_index(drop=True)
        
        # Create pointers for each department
        dept_pointers = {dept: 0 for dept in departments}
        dept_list = list(departments)
        
        # Allocate students to halls
        current_hall_idx = 0
        current_seat_in_hall = 1
        current_dept_idx = 0
        
        allocations = []
        
        total_allocated = 0
        total_students = len(self.students_df)
        
        while total_allocated < total_students:
            # Get current hall info
            hall_no = self.halls_df.loc[current_hall_idx, 'hallno']
            hall_capacity = self.halls_df.loc[current_hall_idx, 'capacity']
            
            # Try to allocate from current department
            dept = dept_list[current_dept_idx]
            dept_pointer = dept_pointers[dept]
            
            # Check if current department still has students
            if dept_pointer < len(dept_groups[dept]):
                student = dept_groups[dept].iloc[dept_pointer]
                
                allocations.append({
                    'Hall No': hall_no,
                    'Seat No': current_seat_in_hall,
                    'Register Number': student['Register Number'],
                    'Name': student['Name'],
                    'Department': student['Department']
                })
                
                dept_pointers[dept] += 1
                current_seat_in_hall += 1
                total_allocated += 1
                
            # Move to next department in rotation
            current_dept_idx = (current_dept_idx + 1) % len(dept_list)
            
            # Check if we need to move to next hall
            if current_seat_in_hall > hall_capacity:
                current_hall_idx += 1
                current_seat_in_hall = 1
                
                if current_hall_idx >= len(self.halls_df):
                    print("Warning: Ran out of halls!")
                    break
        
        allocations_df = pd.DataFrame(allocations)
        print(f"\nTotal students allocated: {total_allocated}")
        print(f"Halls used: {current_hall_idx + 1} out of {len(self.halls_df)}")
        
        return allocations_df
    
    def allocate_seats_same_department(self):
        """
        Allocate seats with same department students together
        Students from same department fill up halls before moving to next department
        """
        print("\n" + "=" * 60)
        print("SEATING ALLOCATION - SAME DEPARTMENT FORMAT")
        print("=" * 60)
        
        # Sort students by department and register number
        students_sorted = self.students_df.sort_values(
            ['Department', 'Register Number']
        ).reset_index(drop=True)
        
        allocations = []
        current_hall_idx = 0
        current_seat_in_hall = 1
        
        for idx, student in students_sorted.iterrows():
            hall_no = self.halls_df.loc[current_hall_idx, 'hallno']
            hall_capacity = self.halls_df.loc[current_hall_idx, 'capacity']
            
            allocations.append({
                'Hall No': hall_no,
                'Seat No': current_seat_in_hall,
                'Register Number': student['Register Number'],
                'Name': student['Name'],
                'Department': student['Department']
            })
            
            current_seat_in_hall += 1
            
            # Move to next hall if current hall is full
            if current_seat_in_hall > hall_capacity:
                current_hall_idx += 1
                current_seat_in_hall = 1
        
        allocations_df = pd.DataFrame(allocations)
        print(f"\nTotal students allocated: {len(allocations_df)}")
        print(f"Halls used: {current_hall_idx + 1} out of {len(self.halls_df)}")
        
        return allocations_df
    
    def _create_hall_wise_summary(self):
        """Create a summary of allocations by hall"""
        self.hall_wise_allocations = {}
        
        for hall_no in self.allocations['Hall No'].unique():
            hall_data = self.allocations[self.allocations['Hall No'] == hall_no].copy()
            hall_data = hall_data.sort_values('Seat No').reset_index(drop=True)
            self.hall_wise_allocations[hall_no] = hall_data
    
    def assign_teachers(self):
        """Assign teachers to halls (one-to-one assignment)"""
        print("\n" + "=" * 60)
        print("ASSIGNING TEACHERS TO HALLS")
        print("=" * 60)
        
        halls_used = sorted(self.hall_wise_allocations.keys())
        teachers_list = self.teachers_df['Name'].str.strip().tolist()
        
        # One-to-one assignment
        for idx, hall_no in enumerate(halls_used):
            if idx < len(teachers_list):
                self.teacher_assignments[hall_no] = teachers_list[idx]
            else:
                self.teacher_assignments[hall_no] = "To be assigned"
        
        print(f"\nAssigned {len(halls_used)} teachers to {len(halls_used)} halls")
        if len(teachers_list) > len(halls_used):
            print(f"Reserve teachers: {', '.join(teachers_list[len(halls_used):])}")
    
    def convert_to_2d_layout(self, hall_no):
        """Convert student list to 2D grid layout using hall-specific columns"""
        # Get the number of columns for this specific hall
        num_cols = self.halls_df[self.halls_df['hallno'] == hall_no]['Columns'].values[0]
            
        hall_data = self.hall_wise_allocations[hall_no]
        hall_capacity = self.halls_df[self.halls_df['hallno'] == hall_no]['capacity'].values[0]
        
        # Calculate number of rows needed
        num_rows = int(np.ceil(hall_capacity / num_cols))
        
        # Create 2D grid
        layout = []
        students = hall_data['Register Number'].tolist()
        
        for row in range(num_rows):
            row_data = []
            for col in range(num_cols):
                idx = row * num_cols + col
                if idx < len(students):
                    row_data.append(students[idx])
                else:
                    row_data.append("EMPTY")
            layout.append(row_data)
        
        return layout, num_rows, num_cols
    
    def generate_hall_visual(self, hall_no, save_path=None):
        """Generate visual representation of hall layout using matplotlib"""
        layout, num_rows, num_cols = self.convert_to_2d_layout(hall_no)
        
        # Create figure in landscape orientation (11.69 x 8.27 inches = A4 landscape)
        fig, ax = plt.subplots(figsize=(11.69, 8.27))
        ax.axis('off')
        
        # Get hall info
        teacher = self.teacher_assignments.get(hall_no, "TBA")
        hall_capacity = self.halls_df[self.halls_df['hallno'] == hall_no]['capacity'].values[0]
        hall_data = self.hall_wise_allocations[hall_no]
        occupied = len(hall_data)
        
        # Get department breakdown
        dept_counts = hall_data['Department'].value_counts()
        dept_text = "\n".join([f"{dept}({count})" for dept, count in dept_counts.items()])
        
        # Add college header
        fig.text(0.5, 0.96, 'Marri Laxman Reddy Institute of Technology',
                ha='center', fontsize=16, fontweight='bold')
        fig.text(0.5, 0.93, 'Hyderabad - 43',
                ha='center', fontsize=11)
        fig.text(0.5, 0.90, '[An Autonomous Institution]',
                ha='center', fontsize=9, style='italic')
        fig.text(0.5, 0.87, 'SEATING ARRANGEMENT',
                ha='center', fontsize=14, fontweight='bold')
        
        # Add date, session, and hall info
        from datetime import date
        today = date.today().strftime('%d-%m-%Y')
        fig.text(0.1, 0.82, f'Date:{today}', fontsize=10)
        fig.text(0.5, 0.82, f'Session:{self.session}', ha='center', fontsize=10)
        fig.text(0.9, 0.82, f'Hall:{hall_no}', ha='right', fontsize=10)
        
        # Create column headers
        col_headers = [str(i+1) for i in range(num_cols)]  # 1, 2, 3, 4, 5...
        
        # Prepare table data with row numbers
        table_data = [col_headers]
        for row_idx, row in enumerate(layout, 1):
            table_data.append(row)
        
        # Create main seating table
        table = ax.table(cellText=table_data, cellLoc='center', loc='center',
                        bbox=[0.1, 0.25, 0.8, 0.52])
        
        # Style the table
        table.auto_set_font_size(False)
        table.set_fontsize(9)
        table.scale(1, 2)
        
        # Style all cells with borders only (no colors)
        for key, cell in table.get_celld().items():
            cell.set_edgecolor('black')
            cell.set_linewidth(1)
            cell.set_facecolor('white')
            
            # Make header row bold
            if key[0] == 0:
                cell.set_text_props(weight='bold')
            
            # Style empty cells
            if cell.get_text().get_text() == "EMPTY":
                cell.set_text_props(color='lightgray', style='italic')
        
        # Add department breakdown table at bottom
        dept_data = [[dept, count] for dept, count in dept_counts.items()]
        dept_data.insert(0, ['Department', 'Count'])
        dept_data.append(['Total Number of Students:', str(occupied)])
        
        dept_table = ax.table(cellText=dept_data, cellLoc='left', loc='lower center',
                             bbox=[0.1, 0.05, 0.5, 0.15])
        dept_table.auto_set_font_size(False)
        dept_table.set_fontsize(9)
        dept_table.scale(1, 1.5)
        
        # Style department table
        for key, cell in dept_table.get_celld().items():
            cell.set_edgecolor('black')
            cell.set_linewidth(1)
            cell.set_facecolor('white')
            if key[0] == 0 or key[0] == len(dept_data) - 1:
                cell.set_text_props(weight='bold')
        
        plt.tight_layout()
        
        if save_path:
            plt.savefig(save_path, dpi=150, bbox_inches='tight', facecolor='white')
            plt.close()
        else:
            return fig
    
    def generate_student_pdf(self, output_file=None):
        """Generate student PDF with hall layouts"""
        if output_file is None:
            output_file = f'seating_student_{self.generation_date}_{self.session}.pdf'
        
        print("\n" + "=" * 60)
        print("GENERATING STUDENT PDF")
        print("=" * 60)
        
        with PdfPages(output_file) as pdf:
            for hall_no in sorted(self.hall_wise_allocations.keys()):
                print(f"  Creating layout for Hall {hall_no}...")
                fig = self.generate_hall_visual(hall_no)
                pdf.savefig(fig, bbox_inches='tight')
                plt.close(fig)
        
        print(f"\n✓ Student PDF generated: {output_file}")
        return output_file
    
    def generate_faculty_pdf(self, output_file=None):
        """Generate faculty PDF with summary table"""
        if output_file is None:
            output_file = f'seating_faculty_{self.generation_date}_{self.session}.pdf'
        
        print("\n" + "=" * 60)
        print("GENERATING FACULTY PDF")
        print("=" * 60)
        
        # Use portrait orientation for faculty summary
        doc = SimpleDocTemplate(output_file, pagesize=A4,
                               rightMargin=30, leftMargin=30,
                               topMargin=30, bottomMargin=30)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # College Header
        header_style = ParagraphStyle(
            'CollegeHeader',
            parent=styles['Heading1'],
            fontSize=16,
            fontName='Helvetica-Bold',
            alignment=TA_CENTER,
            spaceAfter=6
        )
        
        sub_header_style = ParagraphStyle(
            'SubHeader',
            parent=styles['Normal'],
            fontSize=11,
            alignment=TA_CENTER,
            spaceAfter=4
        )
        
        italic_style = ParagraphStyle(
            'Italic',
            parent=styles['Normal'],
            fontSize=9,
            fontName='Helvetica-Oblique',
            alignment=TA_CENTER,
            spaceAfter=6
        )
        
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=14,
            fontName='Helvetica-Bold',
            alignment=TA_CENTER,
            spaceAfter=12
        )
        
        elements.append(Paragraph("Marri Laxman Reddy Institute of Technology", header_style))
        elements.append(Paragraph("Hyderabad - 43", sub_header_style))
        elements.append(Paragraph("[An Autonomous Institution]", italic_style))
        elements.append(Paragraph("SEATING ARRANGEMENT - FACULTY SUMMARY", title_style))
        elements.append(Paragraph(f"Date: {self.generation_date} | Session: {self.session}", sub_header_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Overall Statistics
        total_students = len(self.allocations)
        halls_used = len(self.hall_wise_allocations)
        total_capacity = sum([self.halls_df[self.halls_df['hallno'] == h]['capacity'].values[0] 
                             for h in self.hall_wise_allocations.keys()])
        utilization = (total_students / total_capacity * 100) if total_capacity > 0 else 0
        
        stats_data = [
            ['Overall Statistics', ''],
            ['Total Students', str(total_students)],
            ['Total Halls Used', f"{halls_used} / {len(self.halls_df)}"],
            ['Total Capacity', str(total_capacity)],
            ['Utilization', f"{utilization:.1f}%"]
        ]
        
        stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
        stats_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(stats_table)
        elements.append(Spacer(1, 0.4*inch))
        
        # Hall Allocation Summary Table
        summary_title = Paragraph("Hall Allocation Summary", styles['Heading2'])
        elements.append(summary_title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Prepare table data
        table_data = [['Hall', 'Cap', 'Occ', 'Empty', 'Departments', 'Dept Count', 'Invigilator']]
        
        for hall_no in sorted(self.hall_wise_allocations.keys()):
            hall_data = self.hall_wise_allocations[hall_no]
            capacity = self.halls_df[self.halls_df['hallno'] == hall_no]['capacity'].values[0]
            occupied = len(hall_data)
            empty = capacity - occupied
            
            # Get department counts
            dept_counts = hall_data['Department'].value_counts()
            depts = ', '.join(dept_counts.index.tolist())
            dept_breakdown = ', '.join([f"{dept}({count})" for dept, count in dept_counts.items()])
            
            teacher = self.teacher_assignments.get(hall_no, "TBA")
            
            table_data.append([
                str(hall_no),
                str(capacity),
                str(occupied),
                str(empty),
                depts,
                dept_breakdown,
                teacher
            ])
        
        # Create table
        col_widths = [0.5*inch, 0.6*inch, 0.6*inch, 0.6*inch, 1.2*inch, 1.5*inch, 1.2*inch]
        summary_table = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        summary_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black)
        ]))
        
        elements.append(summary_table)
        
        # Build PDF
        doc.build(elements)
        
        print(f"\n✓ Faculty PDF generated: {output_file}")
        return output_file
    
    def generate_excel_report(self, output_file='seating_allocation_report.xlsx'):
        """Generate comprehensive Excel report with multiple sheets"""
        
        print("\n" + "=" * 60)
        print("GENERATING EXCEL REPORT")
        print("=" * 60)
        
        writer = pd.ExcelWriter(output_file, engine='openpyxl')
        
        # Sheet 1: Complete allocation list (Linear Department)
        self.allocations.to_excel(writer, sheet_name='Complete Allocation', index=False)
        
        # Sheet 2: Hall-wise breakdown
        hall_summary = []
        for hall_no, hall_data in sorted(self.hall_wise_allocations.items()):
            dept_counts = hall_data['Department'].value_counts()
            hall_summary.append({
                'Hall No': hall_no,
                'Total Students': len(hall_data),
                'Departments': ', '.join([f"{dept}({count})" for dept, count in dept_counts.items()])
            })
        
        pd.DataFrame(hall_summary).to_excel(writer, sheet_name='Hall Summary', index=False)
        
        # Sheet 3: Department-wise summary
        dept_summary = self.allocations.groupby('Department').agg({
            'Register Number': 'count',
            'Hall No': lambda x: f"{x.min()} to {x.max()}"
        }).reset_index()
        dept_summary.columns = ['Department', 'Total Students', 'Hall Range']
        dept_summary.to_excel(writer, sheet_name='Department Summary', index=False)
        
        # Create individual hall sheets
        for hall_no, hall_data in sorted(self.hall_wise_allocations.items()):
            sheet_name = f"Hall {hall_no}"
            hall_data.to_excel(writer, sheet_name=sheet_name, index=False)
        
        writer.close()
        
        # Now format the Excel file
        self._format_excel(output_file)
        
        print(f"\n✓ Excel report generated: {output_file}")
        print(f"✓ Total sheets created: {3 + len(self.hall_wise_allocations)}")
        
        return output_file
    
    def _format_excel(self, file_path):
        """Apply formatting to the Excel file"""
        from openpyxl import load_workbook
        from openpyxl.styles import numbers
        
        wb = load_workbook(file_path)
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Find the Register Number column
            reg_num_col = None
            for idx, cell in enumerate(ws[1], 1):
                if cell.value == 'Register Number':
                    reg_num_col = idx
                    break
            
            # Format headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Format data cells
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for idx, cell in enumerate(row, 1):
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Force Register Number column to be text format
                    if reg_num_col and idx == reg_num_col:
                        cell.number_format = '@'  # Text format
                        # Ensure the value is stored as string
                        if cell.value is not None:
                            cell.value = str(cell.value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)
    
    def print_statistics(self):
        """Print allocation statistics"""
        print("\n" + "=" * 60)
        print("ALLOCATION STATISTICS")
        print("=" * 60)
        
        print("\nDepartment-wise allocation:")
        dept_stats = self.allocations.groupby('Department').agg({
            'Register Number': 'count',
            'Hall No': ['min', 'max']
        })
        
        for dept in dept_stats.index:
            count = dept_stats.loc[dept, ('Register Number', 'count')]
            hall_min = dept_stats.loc[dept, ('Hall No', 'min')]
            hall_max = dept_stats.loc[dept, ('Hall No', 'max')]
            print(f"  {dept:8s}: {count:3d} students (Halls {hall_min:2d} to {hall_max:2d})")
        
        print("\nHall utilization:")
        for hall_no in sorted(self.allocations['Hall No'].unique()):
            hall_capacity = self.halls_df[self.halls_df['hallno'] == hall_no]['capacity'].values[0]
            allocated = len(self.allocations[self.allocations['Hall No'] == hall_no])
            utilization = (allocated / hall_capacity) * 100
            print(f"  Hall {hall_no:2d}: {allocated:2d}/{hall_capacity:2d} seats ({utilization:5.1f}% utilized)")


def main():
    """Main execution function"""
    print("\n" + "=" * 60)
    print("SEATING ARRANGEMENT ALLOCATION SYSTEM")
    print("First Year Students - Academic Year 2024-25")
    print("=" * 60)
    
    # File paths
    script_dir = os.path.dirname(os.path.abspath(__file__))
    halls_file = os.path.join(script_dir, 'halls.csv')
    students_file = os.path.join(script_dir, 'year1.csv')
    teachers_file = os.path.join(script_dir, 'Teachers.csv')
    
    # Get session from user (default to FN)
    session = input("\nEnter session (FN/AN) [default: FN]: ").strip().upper()
    if session not in ['FN', 'AN']:
        session = 'FN'
    
    output_file = os.path.join(script_dir, 'first_year_seating_allocation.xlsx')
    
    # Create allocation system
    system = SeatingAllocationSystem(halls_file, students_file, teachers_file, session=session)
    
    # Perform allocation (Linear Department Format - main allocation)
    allocations = system.allocate_seats_mixed_department()
    
    # Assign teachers to halls
    system.assign_teachers()
    
    # Generate comprehensive Excel report
    system.generate_excel_report(output_file)
    
    # Generate PDF reports
    student_pdf = system.generate_student_pdf()
    faculty_pdf = system.generate_faculty_pdf()
    
    # Print statistics
    system.print_statistics()
    
    print("\n" + "=" * 60)
    print("ALLOCATION COMPLETE!")
    print("=" * 60)
    print(f"\nGenerated Files:")
    print(f"  1. Excel Report: {output_file}")
    print(f"  2. Student PDF: {student_pdf}")
    print(f"  3. Faculty PDF: {faculty_pdf}")
    print("\nThe Excel file contains:")
    print("  • Complete Allocation - Linear department format")
    print("  • Hall Summary - Overview of each hall")
    print("  • Department Summary - Department-wise statistics")
    print("  • Individual Hall Sheets - Detailed view for each hall")
    print("\nPDF Files contain:")
    print("  • Student PDF: Visual hall layouts (22 pages)")
    print("  • Faculty PDF: Summary table with teacher assignments (1 page)")
    print("\n")


if __name__ == "__main__":
    main()
